import pandas as pd
from rapidfuzz import process, fuzz
from openpyxl.styles import PatternFill, Font, Alignment

# ──────────────────────────────────────────────────────────────────────────────
# CARGA DE DATOS
# ──────────────────────────────────────────────────────────────────────────────

def cargar_datos(ruta_mayor, ruta_estado_cuenta):
    """
    Carga el mayor contable y el estado de cuenta bancario.
    El estado de cuenta puede tener múltiples pestañas (un banco por pestaña).

    Parámetros:
        ruta_mayor         : ruta al archivo mayor_contable.xlsx
        ruta_estado_cuenta : ruta al archivo estado_cuenta.xlsx

    Retorna:
        df_mayor  : DataFrame del mayor contable
        df_bancos : DataFrame unificado de todos los bancos
    """
    df_mayor = pd.read_excel(ruta_mayor)
    df_mayor['Importe en moneda local'] = df_mayor['Importe en moneda local'].abs()

    xl = pd.ExcelFile(ruta_estado_cuenta)
    df_bancos = pd.concat(
        [pd.read_excel(ruta_estado_cuenta, sheet_name=sheet) for sheet in xl.sheet_names],
        ignore_index=True
    )

    return df_mayor, df_bancos


# ──────────────────────────────────────────────────────────────────────────────
# CAPA 1 — CRUCE EXACTO
# ──────────────────────────────────────────────────────────────────────────────

def cruce_exacto(df_mayor, df_bancos):
    """
    Capa 1: merge exacto por referencia e importe.
    Cruza el mayor contra el banco y viceversa.

    Parámetros:
        df_mayor  : DataFrame del mayor contable
        df_bancos : DataFrame del estado de cuenta bancario

    Retorna:
        df_conc_mayor : mayor con columnas del banco donde hubo match
        df_conc_banco : banco con columnas del mayor donde hubo match
    """
    # Mayor vs Banco
    df_conc_mayor = pd.merge(
        df_mayor,
        df_bancos[['Banco', 'Cuenta bancaria', 'Referencia', 'Fecha valor', 'Importe']],
        left_on=['Ref_transaccion', 'Importe en moneda local'],
        right_on=['Referencia', 'Importe'],
        how='left'
    ).rename(columns={
        'Cuenta bancaria': 'Cuenta acreditada',
        'Fecha valor'    : 'Fecha de acreditacion',
        'Importe'        : 'Importe acreditado'
    })

    # Banco vs Mayor
    df_conc_banco = pd.merge(
        df_bancos,
        df_mayor[['Fecha de documento', 'Ref_transaccion', 'Clave_2', 'Importe en moneda local']],
        left_on=['Referencia', 'Importe'],
        right_on=['Ref_transaccion', 'Importe en moneda local'],
        how='left'
    )

    n_mayor = df_conc_mayor['Referencia_y'].count()
    n_banco = df_conc_banco['Clave_2'].count()
    print(f"[Capa 1] Mayor: {n_mayor}/{len(df_conc_mayor)} coincidencias | "
          f"Banco: {n_banco}/{len(df_conc_banco)} coincidencias")

    return df_conc_mayor, df_conc_banco


# ──────────────────────────────────────────────────────────────────────────────
# OBTENER PENDIENTES
# ──────────────────────────────────────────────────────────────────────────────

def obtener_pendientes(df_conc_mayor, df_conc_banco):
    """
    Filtra los registros sin match en ambos DataFrames.
    Se puede llamar después del cruce exacto y después del fuzzy.

    Parámetros:
        df_conc_mayor : DataFrame del mayor conciliado
        df_conc_banco : DataFrame del banco conciliado

    Retorna:
        df_partidas_pendientes : registros del mayor sin match
        df_depositos_sobrantes : registros del banco sin match
    """
    df_partidas_pendientes = df_conc_mayor[df_conc_mayor['Cuenta acreditada'].isna()].copy()
    df_depositos_sobrantes = df_conc_banco[df_conc_banco['Clave_2'].isna()].copy()

    print(f"[Pendientes] Partidas sin match: {len(df_partidas_pendientes)} | "
          f"Depósitos sobrantes: {len(df_depositos_sobrantes)}")

    return df_partidas_pendientes, df_depositos_sobrantes


# ──────────────────────────────────────────────────────────────────────────────
# CAPA 2 — FUZZY MATCHING
# ──────────────────────────────────────────────────────────────────────────────

def cruce_fuzzy(df_partidas_pendientes, df_depositos_sobrantes, umbral=80):
    """
    Capa 2: fuzzy matching sobre los registros sin match del cruce exacto.
    Primero filtra por importe exacto, luego aplica similitud de texto
    sobre la referencia.

    Parámetros:
        df_partidas_pendientes : registros del mayor sin match
        df_depositos_sobrantes : registros del banco sin match
        umbral                 : score mínimo de similitud (0-100), default 80

    Retorna:
        df_fuzzy_matches : DataFrame con las coincidencias encontradas
    """
    df_dep = df_depositos_sobrantes.assign(
        ref_str=df_depositos_sobrantes['Referencia'].astype(str)
    )
    df_pen = df_partidas_pendientes.assign(
        ref_str=df_partidas_pendientes['Ref_transaccion'].astype(str)
    )

    def _buscar_match(row):
        candidatos = df_pen[df_pen['Importe en moneda local'] == row['Importe']]
        if candidatos.empty:
            return None

        match = process.extractOne(
            row['ref_str'], candidatos['ref_str'], scorer=fuzz.partial_ratio
        )
        if not match or match[1] < umbral:
            return None

        pendiente = candidatos[candidatos['ref_str'] == match[0]].iloc[0]
        return {
            **{f'deposito_{c}': row[c] for c in df_depositos_sobrantes.columns},
            **{f'pendiente_{c}': pendiente[c] for c in df_partidas_pendientes.columns
               if c not in ['Referencia_x', 'Importe en moneda local']},
            'fuzzy_score_referencia': match[1]
        }

    resultados = df_dep.apply(_buscar_match, axis=1).dropna().tolist()
    df_fuzzy_matches = pd.DataFrame(resultados) if resultados else pd.DataFrame()

    print(f"[Capa 2] Fuzzy matches encontrados: {len(df_fuzzy_matches)}")

    return df_fuzzy_matches


# ──────────────────────────────────────────────────────────────────────────────
# ACTUALIZAR CONCILIACIÓN CON FUZZY MATCHES
# ──────────────────────────────────────────────────────────────────────────────

def actualizar_conciliacion(df_conc_mayor, df_conc_banco, df_fuzzy_matches):
    """
    Rellena los campos vacíos en df_conc_mayor y df_conc_banco
    con los resultados del fuzzy matching. Marca las filas con fuzzy_match=True.

    Parámetros:
        df_conc_mayor    : DataFrame del mayor conciliado
        df_conc_banco    : DataFrame del banco conciliado
        df_fuzzy_matches : DataFrame con las coincidencias fuzzy

    Retorna:
        df_conc_mayor : actualizado con fuzzy matches
        df_conc_banco : actualizado con fuzzy matches
    """
    if df_fuzzy_matches.empty:
        print("[Actualizar] No hay fuzzy matches para actualizar.")
        return df_conc_mayor, df_conc_banco

    for _, match in df_fuzzy_matches.iterrows():
        # Actualizar mayor
        mask_mayor = df_conc_mayor['Ref_transaccion'] == match['pendiente_Ref_transaccion']
        df_conc_mayor.loc[mask_mayor, 'Cuenta acreditada']     = match['deposito_Cuenta bancaria']
        df_conc_mayor.loc[mask_mayor, 'Fecha de acreditacion'] = match['deposito_Fecha valor']
        df_conc_mayor.loc[mask_mayor, 'Importe acreditado']    = match['deposito_Importe']
        df_conc_mayor.loc[mask_mayor, 'Referencia_y']          = match['deposito_Referencia']
        df_conc_mayor.loc[mask_mayor, 'Banco']                 = match['deposito_Banco']
        df_conc_mayor.loc[mask_mayor, 'fuzzy_match']           = True

        # Actualizar banco
        mask_banco = df_conc_banco['Referencia'] == match['deposito_Referencia']
        df_conc_banco.loc[mask_banco, 'Ref_transaccion']         = match['pendiente_Ref_transaccion']
        df_conc_banco.loc[mask_banco, 'Clave_2']                 = match['pendiente_Clave_2']
        df_conc_banco.loc[mask_banco, 'Importe en moneda local'] = match['deposito_Importe']
        df_conc_banco.loc[mask_banco, 'fuzzy_match']             = True

    print(f"[Actualizar] {len(df_fuzzy_matches)} registros actualizados en mayor y banco.")

    return df_conc_mayor, df_conc_banco


# ──────────────────────────────────────────────────────────────────────────────
# EXPORTAR A EXCEL
# ──────────────────────────────────────────────────────────────────────────────

def exportar_excel(df_conc_mayor, df_conc_banco, df_partidas_pendientes,
                   df_depositos_sobrantes, df_fuzzy_matches, output_path):
    """
    Exporta todos los DataFrames a un archivo Excel con 6 pestañas y formato.
    Las filas con fuzzy_match se resaltan en amarillo.
    La pestaña Resumen incluye conteos e importes por categoría.

    Parámetros:
        df_conc_mayor          : mayor conciliado
        df_conc_banco          : banco conciliado
        df_partidas_pendientes : partidas sin match
        df_depositos_sobrantes : depósitos sin match
        df_fuzzy_matches       : detalle de matches fuzzy
        output_path            : ruta de salida del archivo Excel
    """
    # ── Construir resumen ────────────────────────────────────────────────────
    total_mayor     = len(df_conc_mayor)
    total_banco     = len(df_conc_banco)
    total_imp_mayor = df_conc_mayor['Importe en moneda local'].sum()
    total_imp_banco = df_conc_banco['Importe'].sum()

    exactos_mayor_df = df_conc_mayor[(df_conc_mayor['Referencia_y'].notna()) & (df_conc_mayor['fuzzy_match'] != True)]
    exactos_banco_df = df_conc_banco[(df_conc_banco['Clave_2'].notna())       & (df_conc_banco['fuzzy_match']  != True)]
    fuzzy_mayor_df   = df_conc_mayor[df_conc_mayor['fuzzy_match'] == True]
    fuzzy_banco_df   = df_conc_banco[df_conc_banco['fuzzy_match']  == True]

    exactos_mayor_cnt = len(exactos_mayor_df);  exactos_mayor_imp = exactos_mayor_df['Importe en moneda local'].sum()
    exactos_banco_cnt = len(exactos_banco_df);  exactos_banco_imp = exactos_banco_df['Importe'].sum()
    fuzzy_mayor_cnt   = len(fuzzy_mayor_df);    fuzzy_mayor_imp   = fuzzy_mayor_df['Importe en moneda local'].sum()
    fuzzy_banco_cnt   = len(fuzzy_banco_df);    fuzzy_banco_imp   = fuzzy_banco_df['Importe'].sum()
    pend_cnt          = len(df_partidas_pendientes); pend_imp      = df_partidas_pendientes['Importe en moneda local'].sum()
    sobr_cnt          = len(df_depositos_sobrantes); sobr_imp      = df_depositos_sobrantes['Importe'].sum()

    total_conc_mayor = exactos_mayor_cnt + fuzzy_mayor_cnt
    total_conc_banco = exactos_banco_cnt + fuzzy_banco_cnt
    total_imp_conc_mayor = exactos_mayor_imp + fuzzy_mayor_imp
    total_imp_conc_banco = exactos_banco_imp + fuzzy_banco_imp

    df_resumen = pd.DataFrame([
        {'Categoría': 'Total registros Mayor',           'Registros Mayor': total_mayor,        'Importe Mayor': total_imp_mayor,        'Registros Banco': '',                 'Importe Banco': ''},
        {'Categoría': 'Total registros Banco',           'Registros Mayor': '',                 'Importe Mayor': '',                     'Registros Banco': total_banco,        'Importe Banco': total_imp_banco},
        {'Categoría': '─────────────────',               'Registros Mayor': '',                 'Importe Mayor': '',                     'Registros Banco': '',                 'Importe Banco': ''},
        {'Categoría': 'Coincidencias exactas',           'Registros Mayor': exactos_mayor_cnt,  'Importe Mayor': exactos_mayor_imp,      'Registros Banco': exactos_banco_cnt,  'Importe Banco': exactos_banco_imp},
        {'Categoría': 'Coincidencias fuzzy',             'Registros Mayor': fuzzy_mayor_cnt,    'Importe Mayor': fuzzy_mayor_imp,        'Registros Banco': fuzzy_banco_cnt,    'Importe Banco': fuzzy_banco_imp},
        {'Categoría': 'Total conciliado',                'Registros Mayor': total_conc_mayor,   'Importe Mayor': total_imp_conc_mayor,   'Registros Banco': total_conc_banco,   'Importe Banco': total_imp_conc_banco},
        {'Categoría': '% Conciliado',                    'Registros Mayor': f'{total_conc_mayor/total_mayor:.1%}' if total_mayor else '0%', 'Importe Mayor': f'{total_imp_conc_mayor/total_imp_mayor:.1%}' if total_imp_mayor else '0%', 'Registros Banco': f'{total_conc_banco/total_banco:.1%}' if total_banco else '0%', 'Importe Banco': f'{total_imp_conc_banco/total_imp_banco:.1%}' if total_imp_banco else '0%'},
        {'Categoría': '─────────────────',               'Registros Mayor': '',                 'Importe Mayor': '',                     'Registros Banco': '',                 'Importe Banco': ''},
        {'Categoría': 'Partidas pendientes (sin match)', 'Registros Mayor': pend_cnt,           'Importe Mayor': pend_imp,               'Registros Banco': '',                 'Importe Banco': ''},
        {'Categoría': 'Depósitos sobrantes (sin match)', 'Registros Mayor': '',                 'Importe Mayor': '',                     'Registros Banco': sobr_cnt,           'Importe Banco': sobr_imp},
        {'Categoría': 'Diferencia (Pend - Sobr)',        'Registros Mayor': pend_cnt-sobr_cnt,  'Importe Mayor': round(pend_imp-sobr_imp, 2), 'Registros Banco': '',             'Importe Banco': ''},
    ])

    # ── Escribir Excel ───────────────────────────────────────────────────────
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_resumen.to_excel(writer,             sheet_name='Resumen',             index=False)
        df_conc_mayor.to_excel(writer,          sheet_name='Mayor vs Banco',      index=False)
        df_conc_banco.to_excel(writer,          sheet_name='Banco vs Mayor',      index=False)
        df_partidas_pendientes.to_excel(writer, sheet_name='Partidas Pendientes', index=False)
        df_depositos_sobrantes.to_excel(writer, sheet_name='Depositos Sobrantes', index=False)
        df_fuzzy_matches.to_excel(writer,       sheet_name='Fuzzy Matches',       index=False)

        yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        green  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        orange = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

        # Formato resumen
        ws_resumen = writer.sheets['Resumen']
        for cell in ws_resumen[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        color_map = {2: green, 3: green, 5: green, 6: yellow, 7: green,
                     8: green, 10: orange, 11: orange, 12: orange}
        for row_idx, row in enumerate(ws_resumen.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.fill = color_map.get(row_idx, PatternFill())
                cell.alignment = Alignment(horizontal='center')

        for col, width in [('A', 38), ('B', 18), ('C', 18), ('D', 18), ('E', 18)]:
            ws_resumen.column_dimensions[col].width = width

        # Filas fuzzy en amarillo
        for sheet_name, df in [('Mayor vs Banco', df_conc_mayor), ('Banco vs Mayor', df_conc_banco)]:
            ws = writer.sheets[sheet_name]
            if 'fuzzy_match' in df.columns:
                for row_idx, valor in enumerate(df['fuzzy_match'], start=2):
                    if valor == True:
                        for cell in ws[row_idx]:
                            cell.fill = yellow

    print(f"[Export] Archivo guardado en: {output_path}")
